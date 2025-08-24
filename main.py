from loguru import logger
import os
from crawler import find_xlsx_links_in_html, download_xlsx_file
from openpyxl import load_workbook

# List of datasets to process
DATASETS = [
    {
        "base_url": "https://www.statistics.gr/en/statistics/-/publication/DKT60/-",
        "folder_name": "MCI"
    },
    {
        "base_url": "https://www.statistics.gr/en/statistics/-/publication/SEL05/-",
        "folder_name": "NFG"
    },
    {
        "base_url": "https://www.statistics.gr/en/statistics/-/publication/SJO01/-",
        "folder_name": "LFS"
    },
    {
        "base_url": "https://www.statistics.gr/en/statistics/-/publication/SJO02/-",
        "folder_name": "LFS"
    },
    {
        "base_url": "https://www.statistics.gr/en/statistics/-/publication/SJO03/-",
        "folder_name": "LFS"
    },
    {
        "base_url": "https://www.statistics.gr/en/statistics/-/publication/SEL03/-",
        "folder_name": "EDP"
    },
    {
        "base_url": "https://www.statistics.gr/en/statistics/-/publication/SOP03/-",
        "folder_name": "BLA"
    },
    {
        "base_url": "https://www.statistics.gr/en/statistics/-/publication/DKT63/-",
        "folder_name": "CCI"
    },
    {
        "base_url": "https://www.statistics.gr/en/statistics/-/publication/DKT90/-",
        "folder_name": "HICP"
    }
    # Add more datasets here as needed
    # {
    #     "base_url": "https://www.statistics.gr/en/statistics/-/publication/OTHER/-",
    #     "folder_name": "OTHER"
    # }
]

def convert_xls_to_xlsx(xls_path):
    """
    Converts an .xls file to .xlsx format and returns the new file path.
    """
    try:
        import xlrd
        from openpyxl import Workbook

        logger.info(f"Converting {xls_path} to .xlsx format...")
        book = xlrd.open_workbook(xls_path)
        wb = Workbook()
        for sheet_index in range(book.nsheets):
            sheet = book.sheet_by_index(sheet_index)
            if sheet_index == 0:
                ws = wb.active
                ws.title = sheet.name
            else:
                ws = wb.create_sheet(title=sheet.name)
            for row in range(sheet.nrows):
                ws.append(sheet.row_values(row))
        xlsx_path = os.path.splitext(xls_path)[0] + ".xlsx"
        wb.save(xlsx_path)
        logger.success(f"Converted {xls_path} to {xlsx_path}")
        return xlsx_path
    except Exception as e:
        logger.error(f"Failed to convert {xls_path} to .xlsx: {e}")
        return None

def unhide_and_unprotect_xlsx(file_path):
    try:
        wb = load_workbook(file_path)
        invisible = 0
        protected = 0
        for sheet in wb.worksheets:
            if sheet.sheet_state != "visible":
                sheet.sheet_state = "visible"
                invisible += 1
            if sheet.protection.sheet:
                sheet.protection.sheet = False
                protected += 1
        wb.save(file_path)
        logger.success(f"Processed {file_path}: {invisible} sheet(s) made visible, {protected} sheet(s) unprotected.")
    except Exception as e:
        logger.error(f"Failed to process {file_path}: {e}")

def process_dataset(base_url, folder_name):
    logger.info(f"Processing dataset: {folder_name} from {base_url}")
    os.makedirs(os.path.join("assets", folder_name), exist_ok=True)
    # Use the correct pattern-based function to find xlsx links
    xlsx_links = find_xlsx_links_in_html(base_url, positions=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20])
    downloaded_files = []
    for xlsx_url in xlsx_links:
        file_path = download_xlsx_file(xlsx_url, subfolder=folder_name)
        if file_path:
            logger.success(f"Downloaded: {file_path}")
            downloaded_files.append(file_path)
        else:
            logger.error(f"Failed to download: {xlsx_url}")

    # Convert all .xls to .xlsx (if any, though the pattern only finds xlsx, but keep for completeness)
    xlsx_files = []
    xls_files_to_delete = []
    for file_path in downloaded_files:
        _, ext = os.path.splitext(file_path)
        if ext.lower() == ".xls":
            xlsx_path = convert_xls_to_xlsx(file_path)
            if xlsx_path:
                xlsx_files.append(xlsx_path)
                xls_files_to_delete.append(file_path)
            else:
                logger.error(f"Failed to convert {file_path} to .xlsx")
        elif ext.lower() == ".xlsx":
            xlsx_files.append(file_path)

    # Delete all .xls files after conversion
    for xls_file in xls_files_to_delete:
        try:
            os.remove(xls_file)
            logger.success(f"Deleted original .xls file: {xls_file}")
        except Exception as e:
            logger.error(f"Failed to delete {xls_file}: {e}")

    # Unhide and unprotect all xlsx files
    for xlsx_file in xlsx_files:
        unhide_and_unprotect_xlsx(xlsx_file)

if __name__ == "__main__":
    for dataset in DATASETS:
        process_dataset(dataset["base_url"], dataset["folder_name"])
