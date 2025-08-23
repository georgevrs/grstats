import os
from openpyxl import load_workbook

def unhide_and_unprotect_xlsx(file_path):
    try:
        wb = load_workbook(file_path)
        for sheet in wb.worksheets:
            if sheet.sheet_state != "visible":
                sheet.sheet_state = "visible"
            sheet.protection.sheet = False
        wb.save(file_path)
        print(f"Processed (overwritten): {file_path}")
    except Exception as e:
        print(f"Failed to process {file_path}: {e}")

def convert_xls_to_xlsx(xls_path, xlsx_path):
    try:
        import xlrd
        from openpyxl import Workbook
        book = xlrd.open_workbook(xls_path)
        wb = Workbook()
        for sheet_index in range(book.nsheets):
            sheet = book.sheet_by_index(sheet_index)
            if sheet_index == 0:
                ws = wb.active
                ws.title = sheet.name
            else:
                ws = wb.create_sheet(title=sheet.name)
            for row in range(sheet.nrows):
                ws.append(sheet.row_values(row))
        wb.save(xlsx_path)
        print(f"Converted: {xls_path} -> {xlsx_path}")
        return True
    except Exception as e:
        print(f"Failed to convert {xls_path}: {e}")
        return False

def process_all_excel_files(assets_dir="assets"):
    for root, dirs, files in os.walk(assets_dir):
        for file in files:
            file_path = os.path.join(root, file)
            name, ext = os.path.splitext(file)
            if ext.lower() == ".xlsx":
                unhide_and_unprotect_xlsx(file_path)
            elif ext.lower() == ".xls":
                # Convert to xlsx first, then unhide/unprotect
                xlsx_path = os.path.join(root, f"{name}.xlsx")
                if convert_xls_to_xlsx(file_path, xlsx_path):
                    unhide_and_unprotect_xlsx(xlsx_path)

if __name__ == "__main__":
    process_all_excel_files("assets")
